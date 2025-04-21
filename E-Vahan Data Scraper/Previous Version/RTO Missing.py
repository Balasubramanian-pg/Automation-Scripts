import os
import time
import logging
import warnings
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import TimeoutException, StaleElementReferenceException, NoSuchElementException, ElementNotInteractableException
import re

# Suppress style warnings
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl.styles.stylesheet')

# Configure logging
logging.basicConfig(
    filename='vahan_analyzer.log',
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)

# Hardcoded paths - MODIFY WITH YOUR SPECIFIC PATHS
ROOT_DIR = r"C:\Users\ASUS\OneDrive\Desktop\Work\Vehicle Category"  # Path to your downloaded Excel files
OUTPUT_DIR = r"C:\Users\ASUS\OneDrive\Desktop\Work\Vehicle Category"  # Path to save reports
SCRAPED_DATA_PATH = os.path.join(OUTPUT_DIR, "vahan_states_rtos.xlsx")  # Path to save scraped data
REPORT_PATH = os.path.join(OUTPUT_DIR, "rto_download_status.xlsx")  # Path to save the final report

# Debug mode - set to True for verbose output
DEBUG_MODE = True

# Chrome path for Selenium
CHROME_PATH = r"C:/Program Files (x86)/chrome-win64/chrome.exe"

# Create output directory if it doesn't exist
os.makedirs(OUTPUT_DIR, exist_ok=True)

# ==============================================
# Part 1: Excel File Processing Functions
# ==============================================

def fast_read_title(file_path):
    """Quickly read first cell using openpyxl without loading full file"""
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        sheet = wb.active
        title_value = sheet.cell(row=1, column=1).value
        
        # Handle None values
        if title_value is None:
            return ""
        return str(title_value)
    except Exception as e:
        logging.error(f"Excel read failed: {file_path}: {str(e)}")
        if DEBUG_MODE:
            print(f"Excel read failed: {file_path}: {str(e)}")
        return ""
    finally:
        if 'wb' in locals():
            try:
                wb.close()
            except:
                pass  # Ignore errors on close

def extract_name_from_title(title, filename):
    """Extract just the location code from the title"""
    logging.info(f"Processing title: '{title}'")
    
    # Specific pattern extraction: Get just "Baratang - AN201" from 
    # "Maker Month Wise Data of Baratang - AN201 , Andaman & Nicobar Island (2024)"
    prefix = "Maker Month Wise Data of "
    if "Maker Month Wise Data" in title:
        # Find the start position after "Maker Month Wise Data of "
        # Handle both "Maker Month Wise Data of" and "Maker Month Wise Data  of" (extra space)
        if "Maker Month Wise Data of " in title:
            start = title.find("Maker Month Wise Data of ") + len("Maker Month Wise Data of ")
        else:
            start = title.find("Maker Month Wise Data  of ") + len("Maker Month Wise Data  of ")
        
        # Find the comma after the location code
        end = title.find(",", start)
        if end != -1:
            return title[start:end].strip()
        else:
            # If no comma, try finding the state name in parentheses
            end = title.find(" (", start)
            if end != -1:
                return title[start:end].strip()
            else:
                # If no obvious delimiter, use the rest of the string
                return title[start:].strip()
    
    # Fallback - use original filename
    logging.warning(f"Could not extract name from title: '{title}', using original filename")
    return os.path.splitext(filename)[0]

def process_file(args):
    """Process a single file to extract RTO information"""
    folder, filename = args
    file_path = os.path.join(folder, filename)
    
    try:
        # Read title from Excel file
        title = fast_read_title(file_path)
        
        if not title:
            return None
        
        # Extract just the location code
        location_code = extract_name_from_title(title, filename)
        
        if DEBUG_MODE and location_code:
            logging.info(f"Extracted location: '{location_code}' from file: {filename}")
        
        return location_code
        
    except Exception as e:
        error_msg = f"{file_path} - {str(e)}"
        logging.error(error_msg)
        if DEBUG_MODE:
            print(f"Error: {error_msg}")
        return None

def process_downloaded_files(max_workers=4):
    """Process all downloaded Excel files to extract RTO information"""
    print("\nAnalyzing downloaded Excel files...")
    
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        # Collect all files first
        all_files = []
        for folder, _, files in os.walk(ROOT_DIR):
            excel_files = [f for f in files if f.lower().endswith(('.xlsx', '.xls'))]
            all_files.extend([(folder, f) for f in excel_files])
        
        file_count = len(all_files)
        print(f"Found {file_count} Excel files to analyze")
        
        # Process files with progress updates
        results = list(executor.map(process_file, all_files))
        downloaded_rtos = [r for r in results if r is not None]
        
        print(f"Successfully extracted RTO information from {len(downloaded_rtos)} files")
        return downloaded_rtos

# ==============================================
# Part 2: Vahan Website Scraping Functions
# ==============================================

def setup_driver(chrome_path=None):
    """Set up and return a Chrome webdriver with appropriate options."""
    chrome_options = Options()
    
    if chrome_path:
        chrome_options.binary_location = chrome_path
    
    # Uncomment the line below if you want to run headless
    # chrome_options.add_argument("--headless")
    chrome_options.add_argument("--window-size=1920,1080")
    chrome_options.add_argument("--disable-notifications")
    chrome_options.add_argument("--disable-popup-blocking")
    
    driver = webdriver.Chrome(options=chrome_options)
    return driver

def find_dynamic_elements(driver):
    """Find dynamic elements on the page and return their IDs."""
    print("Scanning for dynamic j_idt elements...")
    
    dynamic_elements = {
        "state_dropdown_id": None,
        "rto_dropdown_id": None,
        "state_dropdown_label_id": None,
        "rto_dropdown_label_id": None
    }
    
    try:
        # Find the state dropdown
        print("Searching for state dropdown...")
        state_dropdowns = driver.find_elements(By.XPATH, "//div[contains(@id, 'j_idt') and contains(@class, 'ui-selectonemenu')]")
        
        for dropdown in state_dropdowns:
            # Try to find label with text containing "States"
            try:
                label = dropdown.find_element(By.XPATH, ".//label[contains(text(), 'States')]")
                dynamic_elements["state_dropdown_id"] = dropdown.get_attribute("id")
                dynamic_elements["state_dropdown_label_id"] = label.get_attribute("id")
                print(f"Found state dropdown: {dynamic_elements['state_dropdown_id']}")
                break
            except NoSuchElementException:
                continue
        
        # Find the RTO dropdown
        print("Searching for RTO dropdown...")
        rto_dropdowns = driver.find_elements(By.XPATH, "//div[@id='selectedRto' or contains(@id, 'j_idt') and contains(@class, 'ui-selectonemenu')]")
        
        for dropdown in rto_dropdowns:
            # Try to find label with text containing "Office"
            try:
                label = dropdown.find_element(By.XPATH, ".//label[contains(text(), 'Office')]")
                dynamic_elements["rto_dropdown_id"] = dropdown.get_attribute("id")
                dynamic_elements["rto_dropdown_label_id"] = label.get_attribute("id")
                print(f"Found RTO dropdown: {dynamic_elements['rto_dropdown_id']}")
                break
            except NoSuchElementException:
                continue
            
        # Fallback for state dropdown if not found
        if not dynamic_elements["state_dropdown_id"]:
            print("State dropdown not found by label text, trying alternative methods...")
            dropdowns = driver.find_elements(By.XPATH, "//div[contains(@class, 'ui-selectonemenu')]")
            for dropdown in dropdowns:
                dropdown_id = dropdown.get_attribute("id")
                if dropdown_id == "selectedRto":
                    continue  # Skip the known RTO dropdown
                try:
                    dropdown_label = dropdown.find_element(By.XPATH, ".//label")
                    label_id = dropdown_label.get_attribute("id")
                    label_text = dropdown_label.text
                    
                    if "State" in label_text:
                        dynamic_elements["state_dropdown_id"] = dropdown_id
                        dynamic_elements["state_dropdown_label_id"] = label_id
                        print(f"Found state dropdown via label text: {dropdown_id}")
                        break
                except Exception:
                    pass
            
        # Fallback for RTO dropdown if not found
        if not dynamic_elements["rto_dropdown_id"]:
            print("RTO dropdown not found by label text, trying alternative methods...")
            try:
                rto_dropdown = driver.find_element(By.ID, "selectedRto")
                rto_label = rto_dropdown.find_element(By.XPATH, ".//label")
                dynamic_elements["rto_dropdown_id"] = "selectedRto"
                dynamic_elements["rto_dropdown_label_id"] = rto_label.get_attribute("id")
                print(f"Found RTO dropdown via ID: selectedRto")
            except NoSuchElementException:
                remaining_dropdowns = [d for d in driver.find_elements(By.XPATH, "//div[contains(@class, 'ui-selectonemenu')]") 
                                      if d.get_attribute("id") != dynamic_elements["state_dropdown_id"]]
                if remaining_dropdowns:
                    dynamic_elements["rto_dropdown_id"] = remaining_dropdowns[0].get_attribute("id")
                    try:
                        rto_label = remaining_dropdowns[0].find_element(By.XPATH, ".//label")
                        dynamic_elements["rto_dropdown_label_id"] = rto_label.get_attribute("id")
                    except Exception:
                        pass
                    print(f"Found potential RTO dropdown via elimination: {dynamic_elements['rto_dropdown_id']}")
    
    except Exception as e:
        print(f"Error finding dynamic elements: {e}")
    
    return dynamic_elements

def scrape_vahan_states_and_rtos(chrome_path=CHROME_PATH):
    """Scrape state and RTO information from the Vahan website."""
    # Skip scraping if data file exists and user chooses to use it
    if os.path.exists(SCRAPED_DATA_PATH):
        choice = input(f"Found existing scraped data at {SCRAPED_DATA_PATH}. Use this file? (y/n): ").lower()
        if choice == 'y':
            print(f"Using existing data from {SCRAPED_DATA_PATH}")
            return pd.read_excel(SCRAPED_DATA_PATH)
    
    # Initialize the Chrome WebDriver
    driver = setup_driver(chrome_path)
    
    # Navigate to the website
    url = "https://vahan.parivahan.gov.in/vahan4dashboard/vahan/view/reportview.xhtml"
    print(f"Navigating to {url}")
    driver.get(url)
    
    # Wait for the page to load
    wait = WebDriverWait(driver, 20)
    print("Waiting for page to load...")
    time.sleep(5)  # Increased initial wait time
    
    # Store all state-RTO pairs
    all_data = []
    
    try:
        # Find dynamic elements
        dynamic_elements = find_dynamic_elements(driver)
        
        # Use the dynamic element IDs
        state_dropdown_id = dynamic_elements["state_dropdown_id"]
        rto_dropdown_id = dynamic_elements["rto_dropdown_id"]
        state_dropdown_label_id = dynamic_elements["state_dropdown_label_id"]
        rto_dropdown_label_id = dynamic_elements["rto_dropdown_label_id"]
        
        if not state_dropdown_id or not rto_dropdown_id:
            print("Critical elements not found. Taking screenshot and saving page source for debugging...")
            driver.save_screenshot(os.path.join(OUTPUT_DIR, "missing_elements_screenshot.png"))
            with open(os.path.join(OUTPUT_DIR, "page_source.html"), "w", encoding="utf-8") as f:
                f.write(driver.page_source)
            raise Exception("Could not find all required dynamic elements")
        
        # First, get all states using JavaScript for reliability
        print("Fetching all states...")
        driver.execute_script(f"document.getElementById('{state_dropdown_label_id}').click();")
        time.sleep(1)
        
        states_data = driver.execute_script(f"""
            var stateItems = document.querySelectorAll('#{state_dropdown_id}_items li');
            var states = [];
            for (var i = 1; i < stateItems.length; i++) {{  // Skip first "All States" option
                states.push({{
                    label: stateItems[i].getAttribute('data-label'),
                    value: stateItems[i].textContent.trim()
                }});
            }}
            return states;
        """)
        
        # Close the dropdown
        driver.execute_script("document.body.click();")
        time.sleep(1)
        
        print(f"Found {len(states_data)} states to process")
        
        # Process each state
        for idx, state_data in enumerate(states_data):
            state_name = state_data['label']
            print(f"Processing state {idx+1}/{len(states_data)}: {state_name}")
            
            # Use JavaScript to select the state - this is more reliable
            driver.execute_script(f"""
                // First, open the dropdown
                document.getElementById('{state_dropdown_label_id}').click();
                // Small delay
                setTimeout(function() {{
                    // Find and click the state option
                    var items = document.querySelectorAll('#{state_dropdown_id}_items li');
                    for (var i = 0; i < items.length; i++) {{
                        if (items[i].getAttribute('data-label') === "{state_name}") {{
                            items[i].click();
                            break;
                        }}
                    }}
                }}, 500);
            """)
            
            # Wait for state selection to take effect and RTO dropdown to update
            time.sleep(3)
            
            # Now get RTOs using JavaScript
            driver.execute_script(f"document.getElementById('{rto_dropdown_label_id}').click();")
            time.sleep(2)
            
            rtos_data = driver.execute_script(f"""
                var rtoItems = document.querySelectorAll('#{rto_dropdown_id}_items li');
                var rtos = [];
                for (var i = 1; i < rtoItems.length; i++) {{  // Skip first "All Offices" option
                    rtos.push({{
                        label: rtoItems[i].getAttribute('data-label'),
                        value: rtoItems[i].textContent.trim()
                    }});
                }}
                return rtos;
            """)
            
            # Close the dropdown
            driver.execute_script("document.body.click();")
            time.sleep(1)
            
            # Add the RTOs to our data collection
            for rto_data in rtos_data:
                rto_name = rto_data['label']
                if DEBUG_MODE:
                    print(f"  RTO: {rto_name}")
                all_data.append({
                    'State': state_name,
                    'RTO': rto_name
                })
            
            # If no RTOs were found, log a warning
            if not rtos_data:
                print(f"  Warning: No RTOs found for state {state_name}")
                
    except Exception as e:
        print(f"An error occurred during scraping: {e}")
        driver.save_screenshot(os.path.join(OUTPUT_DIR, "error_screenshot.png"))
        with open(os.path.join(OUTPUT_DIR, "error_page_source.html"), "w", encoding="utf-8") as f:
            f.write(driver.page_source)
        print("Saved error screenshot and page source for debugging")
    
    finally:
        # Close the browser
        driver.quit()
        
        # Convert the data to a DataFrame
        df = pd.DataFrame(all_data)
        
        # Save the data to an Excel file
        if not df.empty:
            df.to_excel(SCRAPED_DATA_PATH, index=False)
            print(f"Data saved to {SCRAPED_DATA_PATH}")
            print(f"Total records: {len(df)}")
        else:
            print("No data was collected")
        
        return df

# ==============================================
# Part 3: Analysis and Report Generation
# ==============================================

def match_rto_names(rto_codes, all_rtos):
    """Match downloaded RTO codes with official RTO names"""
    matched = []
    unmatched = []
    
    # Convert all RTO names to lowercase for case-insensitive matching
    all_rtos_lower = {rto.lower(): rto for rto in all_rtos}
    
    for code in rto_codes:
        code_lower = code.lower()
        
        # Check for exact match
        if code_lower in all_rtos_lower:
            matched.append(all_rtos_lower[code_lower])
            continue
            
        # Try partial matches
        matched_any = False
        for rto_lower, original_rto in all_rtos_lower.items():
            # Check if the code contains the RTO name or vice versa
            if code_lower in rto_lower or rto_lower in code_lower:
                matched.append(original_rto)
                matched_any = True
                break
                
        if not matched_any:
            unmatched.append(code)
    
    return matched, unmatched

def create_detailed_report(scraped_df, downloaded_rtos):
    """Create a detailed Excel report showing which RTOs are downloaded and which are missing"""
    print("\nGenerating detailed RTO download status report...")
    
    # Create a new workbook
    wb = openpyxl.Workbook()
    summary_sheet = wb.active
    summary_sheet.title = "Summary"
    
    # Define styles
    header_font = Font(bold=True)
    missing_fill = PatternFill(start_color="FFAAAA", end_color="FFAAAA", fill_type="solid")  # Light red
    downloaded_fill = PatternFill(start_color="AAFFAA", end_color="AAFFAA", fill_type="solid")  # Light green
    complete_fill = PatternFill(start_color="AAAAFF", end_color="AAAAFF", fill_type="solid")  # Light blue
    
    # Setup summary sheet headers
    summary_headers = ["State", "Total RTOs", "Downloaded", "Missing", "Percentage Complete", "Status"]
    for col, header in enumerate(summary_headers, 1):
        cell = summary_sheet.cell(row=1, column=col, value=header)
        cell.font = header_font
    
    # Group the scraped data by state
    state_groups = scraped_df.groupby('State')
    
    # Process each state
    summary_row = 2
    all_rtos = scraped_df['RTO'].tolist()
    
    # Convert downloaded RTOs to lowercase for case-insensitive matching
    downloaded_rtos_lower = [rto.lower() for rto in downloaded_rtos if rto]
    
    # Process each state
    for state_name, state_group in state_groups:
        print(f"Processing state: {state_name}")
        
        # Create a sheet for this state
        state_sheet = wb.create_sheet(title=state_name[:31])  # Excel limits sheet names to 31 chars
        
        # Add headers to the state sheet
        state_headers = ["RTO Name", "Status", "Notes"]
        for col, header in enumerate(state_headers, 1):
            cell = state_sheet.cell(row=1, column=col, value=header)
            cell.font = header_font
        
        # Get all RTOs for this state
        state_rtos = state_group['RTO'].tolist()
        state_rtos_lower = [rto.lower() for rto in state_rtos]
        
        # Check which RTOs have been downloaded
        downloaded_count = 0
        state_row = 2
        
        for rto in state_rtos:
            rto_lower = rto.lower()
            
            # Check if this RTO has been downloaded (using case-insensitive matching)
            is_downloaded = False
            matched_name = None
            
            # Check for exact match
            if rto_lower in downloaded_rtos_lower:
                is_downloaded = True
            
            # Check for partial matches if no exact match
            if not is_downloaded:
                for downloaded_rto in downloaded_rtos_lower:
                    # If either contains the other, consider it a match
                    if rto_lower in downloaded_rto or downloaded_rto in rto_lower:
                        is_downloaded = True
                        matched_name = downloaded_rto
                        break
            
            # Write to the state sheet
            state_sheet.cell(row=state_row, column=1, value=rto)
            
            if is_downloaded:
                status_cell = state_sheet.cell(row=state_row, column=2, value="Downloaded")
                status_cell.fill = downloaded_fill
                if matched_name and matched_name != rto_lower:
                    state_sheet.cell(row=state_row, column=3, value=f"Matched with: {downloaded_rtos[downloaded_rtos_lower.index(matched_name)]}")
                downloaded_count += 1
            else:
                status_cell = state_sheet.cell(row=state_row, column=2, value="Missing")
                status_cell.fill = missing_fill
            
            state_row += 1
        
        # Calculate completion percentage
        total_rtos = len(state_rtos)
        percentage = round((downloaded_count / total_rtos * 100), 2) if total_rtos > 0 else 0
        
        # Add to summary sheet
        summary_sheet.cell(row=summary_row, column=1, value=state_name)
        summary_sheet.cell(row=summary_row, column=2, value=total_rtos)
        summary_sheet.cell(row=summary_row, column=3, value=downloaded_count)
        summary_sheet.cell(row=summary_row, column=4, value=total_rtos - downloaded_count)
        summary_sheet.cell(row=summary_row, column=5, value=f"{percentage}%")
        
        # Status column
        status_cell = summary_sheet.cell(row=summary_row, column=6)
        if percentage == 100:
            status_cell.value = f"Complete - All {total_rtos} RTOs downloaded"
            status_cell.fill = complete_fill
        else:
            status_cell.value = f"{total_rtos - downloaded_count} RTOs missing"
            status_cell.fill = missing_fill
        
        summary_row += 1
    
    # Add a sheet for unmatched downloads
    unmatched_sheet = wb.create_sheet(title="Unmatched Downloads")
    unmatched_sheet.cell(row=1, column=1, value="Downloaded RTO Code")
    unmatched_sheet.cell(row=1, column=2, value="Not Matched with Any Official RTO")
    
    # Find downloaded RTOs that didn't match with any official RTO
    all_rtos_lower = [rto.lower() for rto in all_rtos]
    unmatched_row = 2
    
    for downloaded_rto in downloaded_rtos:
        if not downloaded_rto:
            continue
            
        downloaded_lower = downloaded_rto.lower()
        matched = False
        
        # Check exact and partial matches
        for official_rto_lower in all_rtos_lower:
            if downloaded_lower in official_rto_lower or official_rto_lower in downloaded_lower:
                matched = True
                break
                
        if not matched:
            unmatched_sheet.cell(row=unmatched_row, column=1, value=downloaded_rto)
            unmatched_sheet.cell(row=unmatched_row, column=2, value="Not matched with any official RTO")
            unmatched_row += 1
    
    # Adjust column widths
    for sheet in wb:
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column].width = min(adjusted_width, 50)  # Cap width at 50
    
    # Save the workbook
    wb.save(REPORT_PATH)
    print(f"Report saved to {REPORT_PATH}")

# ==============================================
# Main Execution
# ==============================================

def main():
    print("=" * 80)
    print("Vahan RTO Downloader Analysis Tool")
    print("=" * 80)
    
    try:
        # Step 1: Scrape state and RTO data from the website
        scraped_df = scrape_vahan_states_and_rtos()
        
        if scraped_df is None or scraped_df.empty:
            print("No RTO data available. Exiting...")
            return
            
        # Step 2: Process downloaded Excel files
        downloaded_rtos = process_downloaded_files()
        
        # Step 3: Create detailed report
        create_detailed_report(scraped_df, downloaded_rtos)
        
        print("\nProcess completed successfully!")
        
    except Exception as e:
        print(f"An error occurred: {str(e)}")
        logging.error(f"Main execution error: {str(e)}")

if __name__ == "__main__":
    main()
